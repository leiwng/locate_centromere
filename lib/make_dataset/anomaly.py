"""
这个文件的功能是：根据赖医生给的异常数据，制备用于异常检测的单根数据集。
数据集中数据命名为：ID-index-number-a/n.jpg
如G111文件夹下第2组5号染色体图片的右侧那根（异常）命名为：G111-2-4-a.jpg

这个功能由于比较复杂，计划分为以下几步，将各个部分拆解，以防医生的命名规则等发生变化。

① 获取所有图片（20根）的路径 获取图片异位的两号染色体的号。（init）
② 图片分割，返回分割的单根染色体和号。（segment）
③ 重命名并保存（main）

"""

import cv2
import numpy as np
import os
import shutil
import openpyxl
import warnings


class DataMaker(object):
    def __init__(self):
        self.source_path1 = r'C:\Users\27966\Desktop\working\project&data\简单易位\简单易位1-6'
        self.source_path2 = r'C:\Users\27966\Desktop\working\project&data\简单易位\简单易位6-12'
        self.source_path3 = r'C:\Users\27966\Desktop\working\project&data\简单易位\简单易位12-XY'
        self.target_path = r'C:\Users\27966\Desktop\working\project&data\extracted-yiwei1'

        self.imgs_paths_tmp = [[], [], []]
        self.imgs_paths = []  # 一张图中含有20号染色体的图像路径列表
        self.numbers = []  # 与imgs_paths长度相同且一一对应。每个图像中异常的两个号。

    def init(self):
        # 获取图像的路径
        count = -1
        for root_path in [self.source_path1, self.source_path2, self.source_path3]:
            # 对于简单易位1-6、简单易位6-12、简单易位12-XY这三个文件夹分别处理。
            count += 1
            for item in os.listdir(root_path):
                if os.path.isdir(os.path.join(root_path, item)):  # 是简单异位xx的文件夹
                    # 找到文件夹下的图，并添加到imgs_path中。
                    for img_name in os.listdir(os.path.join(root_path, item)):
                        if img_name[-6:] == '-2.jpg':
                            self.imgs_paths_tmp[count].append(os.path.join(root_path, item, img_name))
                            break

            # 获取每个图像对应的号
            # 首先打开标注文件的working sheet
            if root_path == self.source_path1:
                label_path = os.path.join(root_path, '简单易位1-6.xlsx')
                wb = openpyxl.load_workbook(label_path, read_only=True)
                ws = wb['易位']
            elif root_path == self.source_path2:
                label_path = os.path.join(root_path, '简单易位6-12.xlsx')
                wb = openpyxl.load_workbook(label_path, read_only=True)
                ws = wb['Sheet1']
            elif root_path == self.source_path3:
                label_path = os.path.join(root_path, '简单易位12-XY.xlsx')
                wb = openpyxl.load_workbook(label_path, read_only=True)
                ws = wb['Sheet1']
            else:
                raise ValueError('Wrong value encontered in variable $root_path$')

            # 然后找到图像对应的号。
            # NOTE:有可能找不到
            for item in self.imgs_paths_tmp[count]:
                item_value = item
                found_number = False
                caseID = item.split('\\')[-1].split('-')[0]
                for row in range(2, 100):
                    if ws.cell(row=row, column=4).value == caseID:
                        self.numbers.append((self.number_mapper(ws.cell(row=row, column=12).value),
                                             self.number_mapper(ws.cell(row=row, column=13).value)))
                        found_number = True
                        break
                if not found_number:
                    self.imgs_paths_tmp[count].remove(item_value)
        self.imgs_paths = self.imgs_paths_tmp[0] + self.imgs_paths_tmp[1] + self.imgs_paths_tmp[2]

    def number_mapper(self, value):
        if isinstance(value, int):
            return value - 1
        elif isinstance(value, str):
            # 这里医生给的文件很乱，类型不一致
            if value.lower() == 'x':
                return 22
            elif value.lower() == 'y':
                return 23
            else:
                return int(value) - 1
        else:
            raise ValueError('Invalid data type encontoured.')

    def segment(self, img_path, number,
                gray_threshold=233, area_threshold=1600):
        """
        对于常染色体，一张图中有20根；若有性染色体，一张图中只有15根。
        :param area_threshold:
        :param gray_threshold:
        :param img_path:
        :param number: tuple: (number1, number2), number1 is the number of upper chromosomes.
        :return: list of segmented chromosomes and list of numbers (ordered)
        """
        image = cv2.imdecode(np.fromfile(img_path, dtype=np.uint8), cv2.IMREAD_COLOR)

        # 将红色部分（文字、箭头等标注）变成白色
        lower_red = np.array([0, 0, 150])
        upper_red = np.array([100, 100, 255])
        red_mask = cv2.inRange(image, lower_red, upper_red)
        image[red_mask != 0] = [255, 255, 255]

        # 找到所有染色体的轮廓
        gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
        _, thresholded = cv2.threshold(gray, gray_threshold, 255, cv2.THRESH_BINARY_INV)
        contours, _ = cv2.findContours(thresholded, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        # NOTE:可能有图像上的噪声（医生整理时的问题），导致contour数量大于20。
        contours = [contour for contour in contours if cv2.contourArea(contour) >= area_threshold]

        canvas_list = []
        numbers_list = []

        # NOTE:找到的染色体contour无序。这里要将contours排列成从左到右从上到下的。
        for i in range(len(contours) - 1):
            for j in range(i + 1, len(contours)):
                xi, yi, _, _ = cv2.boundingRect(contours[i])
                xj, yj, _, _ = cv2.boundingRect(contours[j])
                if (yi > image.shape[0] // 3 > yj) or (((yi < image.shape[0] // 3 and yj < image.shape[0] // 3) or
                                                        (yi > image.shape[0] // 3 and yj > image.shape[0] // 3)) and
                                                       xi > xj):
                    contours[i], contours[j] = contours[j], contours[i]

        # NOTE:两根染色体可能被划分到一个contour里，导致contour数量不足20。故这里想办法把含有两个染色体的contour删去。
        # NOTE:性染色体易位的情况只有15个contour！
        if not (22 in number or 23 in number):  # 不涉及性染色体的易位
            num_repeat = 20 - len(contours)
            for _ in range(num_repeat):
                max_area_pos = 0
                for i in range(len(contours)):
                    if cv2.contourArea(contours[i]) > cv2.contourArea(contours[max_area_pos]):
                        max_area_pos = i
                contours.pop(max_area_pos)

        # 对每根染色体的contour：
        for i in range(len(contours)):
            # 生成白板，将mask的染色体区域复制到白板的中心区域，并resize回224*224
            mask = np.full_like(image, fill_value=255)
            cv2.drawContours(mask, [contours[i]], 0, (255, 255, 255), thickness=cv2.FILLED)
            x, y, w, h = cv2.boundingRect(contours[i])

            canvas = np.full(shape=(int(image.shape[0] * 2 / 3), int(image.shape[0] * 2 / 3), 3),
                             fill_value=255,
                             dtype=np.uint8)
            center_x = canvas.shape[1] // 2
            center_y = canvas.shape[0] // 2
            start_x = center_x - w // 2
            start_y = center_y - h // 2
            canvas[start_y:start_y + h, start_x:start_x + w] = cv2.bitwise_and(image[y:y + h, x:x + w],
                                                                               mask[y:y + h, x:x + w])
            canvas = cv2.resize(canvas, (224, 224))
            canvas_list.append(canvas)

            # 根据染色体位置（上下）来确定染色体的号
            numbers_list.append(number[0] if y < image.shape[0] // 3 else number[1])

        '''image = cv2.drawContours(image, contours, -1, (255, 0, 0), thickness=3)
        print(len(contours))
        cv2.namedWindow('1', cv2.WINDOW_NORMAL)
        cv2.imshow('1', image)
        cv2.waitKey(0)
        cv2.destroyAllWindows()'''

        return canvas_list, numbers_list

    def create_folder(self, root_path):
        """
        目标：创建文件夹。
        """
        if os.path.exists(root_path):
            shutil.rmtree(root_path)

        os.mkdir(root_path)
        for i in range(24):
            os.mkdir(os.path.join(root_path, str(i)))

    def main(self):
        print('Extracting anomaly info. from files...')
        self.init()
        self.create_folder(self.target_path)
        print('Finish!\n')

        print('Segmenting chromosomes...')
        # 对每张图：
        count = 0
        error_contour = 0
        print(len(self.imgs_paths), len(self.numbers))
        for img_path, number in zip(self.imgs_paths, self.numbers):
            count += 1
            canvas, numbers = self.segment(img_path, number)
            if ((22 not in numbers and 23 not in numbers) and len(canvas) != 20) or ((22 in numbers or 23 in numbers) and len(canvas) != 15):
                error_contour += 1
            print('\r', count, '/', len(self.imgs_paths), '...\t(', error_contour, ' error)', end='')
            # 对每张图中的每根染色体：
            if not (22 in numbers or 23 in numbers):
                for i in range(len(canvas)):
                    # 命名
                    name_id = img_path.split('\\')[-1].split('-')[0]
                    name_index = str((i % 10) // 2)
                    name_number = str(numbers[i])
                    name_anomalyinfo = 'a' if i % 2 else 'n'
                    new_name = name_id + '-' + name_index + '-' + name_number + '-' + name_anomalyinfo + '.jpg'
                    # 存储
                    cv2.imwrite(os.path.join(self.target_path, str(numbers[i]), new_name), canvas[i])
            else:
                sex_num = numbers.count(22) + numbers.count(23)
                for i in range(len(canvas)):
                    # 命名
                    name_id = img_path.split('\\')[-1].split('-')[0]
                    if i < sex_num:
                        name_index = str(i)
                    else:
                        name_index = str((i - sex_num) // 2)
                    name_number = str(numbers[i])
                    if numbers[i] == 23:
                        name_anomalyinfo = 'a'
                    elif numbers[i] == 22:
                        if sex_num <= 5:
                            name_anomalyinfo = 'a'
                        else:
                            name_anomalyinfo = 'n' if i % 2 == 0 else 'a'
                    else:
                        name_anomalyinfo = 'n' if i % 2 == sex_num % 2 else 'a'
                    new_name = name_id + '-' + name_index + '-' + name_number + '-' + name_anomalyinfo + '.jpg'
                    # 存储
                    cv2.imwrite(os.path.join(self.target_path, str(numbers[i]), new_name), canvas[i])

        print('\nFinish!\n')


if __name__ == '__main__':
    m = DataMaker()
    m.main()
